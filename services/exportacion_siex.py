from datetime import date, datetime
from io import BytesIO
import re
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from core.actuaciones_multicultivo import leer_detalles as _leer_detalles_actuacion
from core.config import APP_NAME
from core.db import conectar


AVISO_NO_OFICIAL = (
    "Este Excel no es un formato oficial SIEX/CUE. Es una exportación "
    "asistida preliminar para revisión del agricultor, asesor o entidad "
    "autorizada. La carga oficial deberá realizarse por los canales oficiales "
    "correspondientes."
)

MIME_XLSX = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

COLUMNAS_VALIDACION = [
    "area",
    "registro_id",
    "gravedad",
    "campo",
    "problema",
    "recomendacion",
    "bloquea_exportacion",
]

COLUMNAS_EXPLOTACION = [
    "explotacion_id",
    "nombre_explotacion",
    "titular",
    "nif",
    "direccion",
    "municipio",
    "provincia",
    "telefono",
    "email",
    "identificador_REA_REGEA_REGEPA",
    "responsable",
    "asesor",
    "numero_asesor",
    "observaciones",
]

COLUMNAS_PARCELAS = [
    "parcela_id",
    "nombre",
    "campaña",
    "provincia_sigpac",
    "municipio_sigpac",
    "agregado_sigpac",
    "zona_sigpac",
    "poligono",
    "parcela",
    "recinto",
    "superficie_sigpac",
    "cultivo_asociado",
    "geometria_disponible",
    "observaciones",
]

COLUMNAS_CULTIVOS = [
    "cultivo_id",
    "campaña",
    "cultivo",
    "variedad",
    "año_plantacion",
    "superficie",
    "parcelas_asociadas",
    "codigo_cultivo_siex",
]

COLUMNAS_TRATAMIENTOS = [
    "tratamiento_id",
    "campaña",
    "fecha_inicio",
    "fecha_fin",
    "cultivo",
    "parcelas",
    "producto",
    "numero_registro_producto",
    "materia_activa",
    "plaga_motivo",
    "dosis",
    "caldo",
    "superficie_tratada",
    "aplicador",
    "equipo_aplicacion",
    "eficacia",
    "plazo_seguridad",
    "recetas_pdf",
    "observaciones",
]

COLUMNAS_FERTILIZACION = [
    "fertilizacion_id",
    "campaña",
    "fecha",
    "cultivo",
    "parcelas",
    "producto",
    "tipo_fertilizante",
    "cantidad",
    "unidad",
    "superficie",
    "codigo_actuacion_siex",
    "unidad_normalizada",
    "observaciones",
]

COLUMNAS_PRACTICAS = [
    "practica_id",
    "campaña",
    "fecha",
    "labor",
    "cultivo",
    "parcelas",
    "superficie",
    "maquinaria",
    "prestador",
    "codigo_actuacion_siex",
    "observaciones",
]

COLUMNAS_COSECHA = [
    "cosecha_id",
    "campaña",
    "fecha",
    "cultivo",
    "parcelas",
    "superficie",
    "cantidad",
    "unidad",
    "destino",
    "cliente",
    "observaciones",
]

COLUMNAS_MAQUINARIA = [
    "id_visual",
    "origen",
    "tipo",
    "marca",
    "modelo",
    "matricula",
    "numero_roma",
    "descripcion",
    "observaciones",
]

COLUMNAS_DOCUMENTOS = [
    "documento_id",
    "tipo_documento",
    "area",
    "registro_id",
    "nombre_original",
    "ruta_relativa",
    "sha256",
    "size_bytes",
    "observaciones",
]


def _texto(valor):

    if valor is None:

        return ""

    try:

        if pd.isna(valor):

            return ""

    except (TypeError, ValueError):

        pass

    return str(valor).strip()


def _int_o_none(valor):

    numero = pd.to_numeric(valor, errors="coerce")

    if pd.isna(numero):

        return None

    return int(numero)


def _numero_o_vacio(valor):

    if _texto(valor) == "":

        return ""

    numero = pd.to_numeric(valor, errors="coerce")

    if pd.isna(numero):

        return valor

    return float(numero)


def _tabla_existe(conn, tabla):

    try:

        fila = conn.execute(
            """
            SELECT 1
            FROM sqlite_master
            WHERE type='table'
            AND name=?
            """,
            (tabla,),
        ).fetchone()

    except Exception:

        return False

    return fila is not None


def _leer_tabla(conn, tabla):

    if not _tabla_existe(conn, tabla):

        return pd.DataFrame()

    try:

        return pd.read_sql_query(f'SELECT * FROM "{tabla}"', conn)

    except Exception:

        return pd.DataFrame()


def _dataframe(columnas, filas=None):

    return pd.DataFrame(filas or [], columns=columnas)


def _filtrar_campana(dataframe, campana_id):

    if (
        dataframe.empty
        or campana_id is None
        or "campana_id" not in dataframe.columns
    ):

        return dataframe.copy()

    campanas = pd.to_numeric(dataframe["campana_id"], errors="coerce")
    return dataframe[campanas == int(campana_id)].copy()


def _mapa_por_id(dataframe):

    if dataframe.empty or "id" not in dataframe.columns:

        return {}

    resultado = {}

    for _, fila in dataframe.iterrows():

        fila_id = _int_o_none(fila.get("id"))

        if fila_id is not None:

            resultado[fila_id] = fila.to_dict()

    return resultado


def _campana_nombre(conn, campana_id):

    if campana_id is None:

        return ""

    campanas = _leer_tabla(conn, "campanas")

    if campanas.empty or "id" not in campanas.columns:

        return str(campana_id)

    ids = pd.to_numeric(campanas["id"], errors="coerce")
    coincidentes = campanas[ids == int(campana_id)]

    if coincidentes.empty:

        return str(campana_id)

    return _texto(coincidentes.iloc[0].get("nombre")) or str(campana_id)


def _campana_nombre_por_id(campanas, campana_id):

    if campanas.empty or campana_id is None or "id" not in campanas.columns:

        return ""

    ids = pd.to_numeric(campanas["id"], errors="coerce")
    coincidentes = campanas[ids == int(campana_id)]

    if coincidentes.empty:

        return ""

    return _texto(coincidentes.iloc[0].get("nombre"))


def _etiqueta_parcela(fila):

    if fila is None:

        return ""

    nombre = _texto(fila.get("nombre"))
    referencia = "-".join(
        parte
        for parte in [
            _texto(fila.get("poligono")),
            _texto(fila.get("parcela")),
            _texto(fila.get("recinto")),
        ]
        if parte
    )

    return nombre or referencia or (
        f"Parcela {_texto(fila.get('id'))}"
        if _texto(fila.get("id"))
        else ""
    )


def _etiqueta_cultivo(fila):

    if fila is None:

        return ""

    partes = [
        _texto(fila.get("especie")) or _texto(fila.get("nombre")),
        _texto(fila.get("variedad")),
        _texto(fila.get("sistema")),
    ]
    ano_plantacion = _texto(fila.get("ano_plantacion"))

    if ano_plantacion:

        partes.append(f"Plantación {ano_plantacion}")

    texto = " / ".join(parte for parte in partes if parte)
    return texto or (
        f"Cultivo {_texto(fila.get('id'))}"
        if _texto(fila.get("id"))
        else ""
    )


def _descripcion_maquinaria(fila):

    if fila is None:

        return ""

    partes = [
        _texto(fila.get("nombre")),
        _texto(fila.get("marca")),
        _texto(fila.get("modelo")),
        _texto(fila.get("tipo")),
    ]
    return " / ".join(parte for parte in partes if parte)


def _agrupar_parcelas(relaciones, campo_id, parcelas_por_id):

    resultado = {}

    if (
        relaciones.empty
        or campo_id not in relaciones.columns
        or "parcela_id" not in relaciones.columns
    ):

        return resultado

    for _, fila in relaciones.iterrows():

        registro_id = _int_o_none(fila.get(campo_id))
        parcela_id = _int_o_none(fila.get("parcela_id"))

        if registro_id is None or parcela_id is None:

            continue

        etiqueta = _etiqueta_parcela(parcelas_por_id.get(parcela_id))

        if etiqueta:

            resultado.setdefault(registro_id, []).append(etiqueta)

    return {
        registro_id: ", ".join(dict.fromkeys(etiquetas))
        for registro_id, etiquetas in resultado.items()
    }


def _agrupar_documentos(documentos, campo_id):

    resultado = {}

    if documentos.empty or campo_id not in documentos.columns:

        return resultado

    for _, fila in documentos.iterrows():

        registro_id = _int_o_none(fila.get(campo_id))

        if registro_id is None:

            continue

        nombre = _texto(fila.get("nombre_original")) or _texto(
            fila.get("ruta_relativa")
        )
        resultado.setdefault(registro_id, []).append(nombre or "PDF")

    return {
        registro_id: ", ".join(dict.fromkeys(nombres))
        for registro_id, nombres in resultado.items()
    }


def _normalizar_eficacia(valor):

    texto = _texto(valor).upper()

    if texto in {"BUENA", "BUENO"}:

        return "B"

    if texto == "REGULAR":

        return "R"

    if texto in {"MALA", "MALO"}:

        return "M"

    return texto if texto in {"B", "R", "M"} else ""


def _sanear_nombre_archivo(texto):

    texto = unicodedata.normalize("NFKD", _texto(texto))
    texto = "".join(
        caracter
        for caracter in texto
        if not unicodedata.combining(caracter)
    )
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", texto)
    texto = re.sub(r"_+", "_", texto).strip("_")
    return texto or "sin_campana"


def _preparar_excel(dataframe):

    preparado = dataframe.copy()

    for columna in preparado.columns:

        if (
            pd.api.types.is_object_dtype(preparado[columna].dtype)
            or pd.api.types.is_string_dtype(preparado[columna].dtype)
        ):

            preparado[columna] = preparado[columna].map(
                lambda valor: (
                    "'" + valor
                    if isinstance(valor, str)
                    and valor.startswith(("=", "+", "-", "@"))
                    else valor
                )
            )

    return preparado.fillna("")


def obtener_dataframe_explotacion(conn):

    explotacion = _leer_tabla(conn, "explotacion")

    if explotacion.empty:

        return _dataframe(COLUMNAS_EXPLOTACION)

    filas = []

    for _, fila in explotacion.iterrows():

        identificador = next(
            (
                _texto(fila.get(campo))
                for campo in [
                    "registro_explotacion",
                    "codigo_regea",
                    "codigo_regepa",
                ]
                if _texto(fila.get(campo))
            ),
            "",
        )
        responsable = " / ".join(
            parte
            for parte in [
                _texto(fila.get("responsable_nombre")),
                _texto(fila.get("responsable_nif")),
                _texto(fila.get("responsable_telefono")),
            ]
            if parte
        )
        asesor = " / ".join(
            parte
            for parte in [
                _texto(fila.get("asesor_nombre")),
                _texto(fila.get("asesor_nif")),
                _texto(fila.get("asesor_telefono")),
            ]
            if parte
        )
        filas.append({
            "explotacion_id": fila.get("id", ""),
            "nombre_explotacion": fila.get("nombre_explotacion", ""),
            "titular": fila.get("titular", ""),
            "nif": fila.get("nif", ""),
            "direccion": fila.get("direccion", ""),
            "municipio": fila.get("localidad", ""),
            "provincia": fila.get("provincia", ""),
            "telefono": fila.get("telefono", ""),
            "email": fila.get("email", ""),
            "identificador_REA_REGEA_REGEPA": identificador,
            "responsable": responsable,
            "asesor": asesor,
            "numero_asesor": fila.get("asesor_numero_registro", ""),
            "observaciones": fila.get("observaciones", ""),
        })

    return _dataframe(COLUMNAS_EXPLOTACION, filas)


def obtener_dataframe_parcelas(conn, campana_nombre=""):

    parcelas = _leer_tabla(conn, "parcelas")

    if parcelas.empty:

        return _dataframe(COLUMNAS_PARCELAS)

    cultivos = _leer_tabla(conn, "cultivos")
    cultivos_por_parcela = {}

    if not cultivos.empty and "parcela_id" in cultivos.columns:

        for _, cultivo in cultivos.iterrows():

            parcela_id = _int_o_none(cultivo.get("parcela_id"))

            if parcela_id is None:

                continue

            etiqueta = _etiqueta_cultivo(cultivo.to_dict())

            if etiqueta:

                cultivos_por_parcela.setdefault(parcela_id, []).append(etiqueta)

    cultivo_parcelas = _leer_tabla(conn, "cultivo_parcelas")

    if (
        not cultivos.empty
        and not cultivo_parcelas.empty
        and "cultivo_id" in cultivo_parcelas.columns
        and "parcela_id" in cultivo_parcelas.columns
    ):

        cultivos_por_id = _mapa_por_id(cultivos)

        for _, relacion in cultivo_parcelas.iterrows():

            parcela_id = _int_o_none(relacion.get("parcela_id"))
            cultivo_id = _int_o_none(relacion.get("cultivo_id"))

            if parcela_id is None or cultivo_id is None:

                continue

            etiqueta = _etiqueta_cultivo(cultivos_por_id.get(cultivo_id))

            if etiqueta:

                cultivos_por_parcela.setdefault(parcela_id, []).append(etiqueta)

    filas = []

    for _, fila in parcelas.iterrows():

        parcela_id = _int_o_none(fila.get("id"))
        geometria = any(
            _texto(fila.get(campo))
            for campo in ["geometry", "sigpac_geojson"]
        )
        cultivo_asociado = ", ".join(
            dict.fromkeys(cultivos_por_parcela.get(parcela_id, []))
        )
        filas.append({
            "parcela_id": fila.get("id", ""),
            "nombre": fila.get("nombre", ""),
            "campaña": campana_nombre,
            "provincia_sigpac": fila.get("provincia_sigpac", ""),
            "municipio_sigpac": fila.get("municipio_sigpac", ""),
            "agregado_sigpac": fila.get("agregado_sigpac", ""),
            "zona_sigpac": fila.get("zona_sigpac", ""),
            "poligono": fila.get("poligono", ""),
            "parcela": fila.get("parcela", ""),
            "recinto": fila.get("recinto", ""),
            "superficie_sigpac": fila.get("superficie_sigpac", ""),
            "cultivo_asociado": cultivo_asociado,
            "geometria_disponible": "Sí" if geometria else "No",
            "observaciones": fila.get("observaciones", ""),
        })

    return _dataframe(COLUMNAS_PARCELAS, filas)


def obtener_dataframe_cultivos(conn, campana_id=None):

    cultivos = _filtrar_campana(_leer_tabla(conn, "cultivos"), campana_id)

    if cultivos.empty:

        return _dataframe(COLUMNAS_CULTIVOS)

    parcelas = _leer_tabla(conn, "parcelas")
    parcelas_por_id = _mapa_por_id(parcelas)
    cultivo_parcelas = _leer_tabla(conn, "cultivo_parcelas")
    parcelas_por_cultivo = _agrupar_parcelas(
        cultivo_parcelas,
        "cultivo_id",
        parcelas_por_id,
    )
    campanas = _leer_tabla(conn, "campanas")
    filas = []

    for _, fila in cultivos.iterrows():

        cultivo_campana = ""

        if "campana_id" in cultivos.columns:

            cultivo_campana = _campana_nombre_por_id(
                campanas,
                _int_o_none(fila.get("campana_id")),
            )

        superficie = fila.get("superficie", "") if "superficie" in cultivos else ""
        parcela = parcelas_por_id.get(_int_o_none(fila.get("parcela_id")))
        parcelas_asociadas = parcelas_por_cultivo.get(
            _int_o_none(fila.get("id")),
            "",
        ) or _etiqueta_parcela(parcela)
        filas.append({
            "cultivo_id": fila.get("id", ""),
            "campaña": cultivo_campana,
            "cultivo": _etiqueta_cultivo(fila.to_dict()),
            "variedad": fila.get("variedad", ""),
            "año_plantacion": fila.get("ano_plantacion", ""),
            "superficie": superficie,
            "parcelas_asociadas": parcelas_asociadas,
            "codigo_cultivo_siex": fila.get("codigo_siex", ""),
        })

    return _dataframe(COLUMNAS_CULTIVOS, filas)


def obtener_dataframe_tratamientos(conn, campana_id=None):

    tratamientos = _filtrar_campana(_leer_tabla(conn, "tratamientos"), campana_id)

    if tratamientos.empty:

        return _dataframe(COLUMNAS_TRATAMIENTOS)

    campanas = _leer_tabla(conn, "campanas")
    cultivos = _mapa_por_id(_leer_tabla(conn, "cultivos"))
    productos = _mapa_por_id(_leer_tabla(conn, "productos_fito"))
    personas = _mapa_por_id(_leer_tabla(conn, "personas"))
    equipos = _mapa_por_id(_leer_tabla(conn, "equipos_aplicacion"))
    maquinaria = _mapa_por_id(_leer_tabla(conn, "maquinaria"))
    parcelas = _mapa_por_id(_leer_tabla(conn, "parcelas"))
    parcelas_por_tratamiento = _agrupar_parcelas(
        _leer_tabla(conn, "tratamiento_parcelas"),
        "tratamiento_id",
        parcelas,
    )
    detalles = _leer_detalles_actuacion(
        conn,
        "tratamiento_cultivos",
        "tratamiento_id",
    )
    documentos = _leer_tabla(conn, "tratamientos_documentos")

    if not documentos.empty and "tipo_documento" in documentos.columns:

        documentos = documentos[
            documentos["tipo_documento"].fillna("").astype(str) == "receta"
        ].copy()

    recetas_por_tratamiento = _agrupar_documentos(
        documentos,
        "tratamiento_id",
    )
    filas = []

    for _, fila in tratamientos.iterrows():

        tratamiento_id = _int_o_none(fila.get("id"))
        campana = _campana_nombre_por_id(
            campanas,
            _int_o_none(fila.get("campana_id")),
        )
        producto = productos.get(_int_o_none(fila.get("producto_id")), {})
        cultivo = cultivos.get(_int_o_none(fila.get("cultivo_id")), {})
        aplicador = personas.get(_int_o_none(fila.get("aplicador_id")), {})
        equipo = equipos.get(
            _int_o_none(fila.get("equipo_aplicacion_id"))
            or _int_o_none(fila.get("equipo_id")),
            {},
        )
        maquinaria_item = maquinaria.get(_int_o_none(fila.get("maquinaria_id")), {})
        equipo_texto = _descripcion_maquinaria(equipo) or _descripcion_maquinaria(
            maquinaria_item
        )
        aplicador_texto = _texto(aplicador.get("nombre")) or _texto(
            fila.get("aplicador")
        )
        fila_base = {
            "tratamiento_id": fila.get("id", ""),
            "campaña": campana,
            "fecha_inicio": fila.get("fecha_inicio", "") or fila.get("fecha", ""),
            "fecha_fin": fila.get("fecha_fin", "") or fila.get("fecha", ""),
            "cultivo": _etiqueta_cultivo(cultivo),
            "parcelas": parcelas_por_tratamiento.get(tratamiento_id, ""),
            "producto": producto.get("nombre", ""),
            "numero_registro_producto": (
                producto.get("registro", "")
                or producto.get("numero_registro", "")
            ),
            "materia_activa": producto.get("materia_activa", ""),
            "plaga_motivo": (
                fila.get("plaga_motivo", "")
                or fila.get("plaga", "")
                or fila.get("problema", "")
                or fila.get("justificacion", "")
            ),
            "dosis": fila.get("dosis", ""),
            "caldo": fila.get("caldo", ""),
            "superficie_tratada": fila.get("superficie_tratada", ""),
            "aplicador": aplicador_texto,
            "equipo_aplicacion": equipo_texto,
            "eficacia": _normalizar_eficacia(fila.get("eficacia")),
            "plazo_seguridad": (
                fila.get("plazo_seguridad", "")
                or producto.get("plazo_seguridad", "")
            ),
            "recetas_pdf": recetas_por_tratamiento.get(tratamiento_id, ""),
            "observaciones": fila.get("observaciones", ""),
        }
        detalles_tratamiento = (
            detalles[
                pd.to_numeric(detalles["registro_id"], errors="coerce")
                == tratamiento_id
            ].copy()
            if not detalles.empty
            and tratamiento_id is not None
            else pd.DataFrame()
        )

        if not detalles_tratamiento.empty:

            for _, detalle in detalles_tratamiento.iterrows():

                cultivo_detalle = cultivos.get(
                    _int_o_none(detalle.get("cultivo_id")),
                    cultivo,
                )
                fila_detalle = fila_base.copy()
                fila_detalle["cultivo"] = (
                    _etiqueta_cultivo(cultivo_detalle)
                    or detalle.get("cultivo", "")
                    or fila_base["cultivo"]
                )
                fila_detalle["parcelas"] = detalle.get("parcela", "")

                if _texto(detalle.get("superficie")):

                    fila_detalle["superficie_tratada"] = detalle.get(
                        "superficie"
                    )

                filas.append(fila_detalle)

            continue

        filas.append(fila_base)

    return _dataframe(COLUMNAS_TRATAMIENTOS, filas)


def obtener_dataframe_fertilizacion(conn, campana_id=None):

    fertilizaciones = _filtrar_campana(
        _leer_tabla(conn, "fertilizaciones"),
        campana_id,
    )

    if fertilizaciones.empty:

        return _dataframe(COLUMNAS_FERTILIZACION)

    campanas = _leer_tabla(conn, "campanas")
    cultivos = _mapa_por_id(_leer_tabla(conn, "cultivos"))
    parcelas = _mapa_por_id(_leer_tabla(conn, "parcelas"))
    parcelas_por_fertilizacion = _agrupar_parcelas(
        _leer_tabla(conn, "fertilizacion_parcelas"),
        "fertilizacion_id",
        parcelas,
    )
    detalles = _leer_detalles_actuacion(
        conn,
        "fertilizacion_cultivos",
        "fertilizacion_id",
    )
    filas = []

    for _, fila in fertilizaciones.iterrows():

        fertilizacion_id = _int_o_none(fila.get("id"))
        cultivo = cultivos.get(_int_o_none(fila.get("cultivo_id")), {})
        fila_base = {
            "fertilizacion_id": fila.get("id", ""),
            "campaña": _campana_nombre_por_id(
                campanas,
                _int_o_none(fila.get("campana_id")),
            ),
            "fecha": fila.get("fecha", ""),
            "cultivo": _etiqueta_cultivo(cultivo) or fila.get("cultivo", ""),
            "parcelas": parcelas_por_fertilizacion.get(fertilizacion_id, ""),
            "producto": fila.get("producto", ""),
            "tipo_fertilizante": (
                fila.get("tipo_fertilizante", "") or fila.get("tipo", "")
            ),
            "cantidad": fila.get("cantidad", ""),
            "unidad": fila.get("unidad", ""),
            "superficie": fila.get("superficie", ""),
            "codigo_actuacion_siex": fila.get("codigo_actuacion_siex", ""),
            "unidad_normalizada": fila.get("unidad_normalizada", ""),
            "observaciones": fila.get("observaciones", ""),
        }
        detalles_fertilizacion = (
            detalles[
                pd.to_numeric(detalles["registro_id"], errors="coerce")
                == fertilizacion_id
            ].copy()
            if not detalles.empty
            and fertilizacion_id is not None
            else pd.DataFrame()
        )

        if not detalles_fertilizacion.empty:

            for _, detalle in detalles_fertilizacion.iterrows():

                cultivo_detalle = cultivos.get(
                    _int_o_none(detalle.get("cultivo_id")),
                    cultivo,
                )
                fila_detalle = fila_base.copy()
                fila_detalle["cultivo"] = (
                    _etiqueta_cultivo(cultivo_detalle)
                    or detalle.get("cultivo", "")
                    or fila_base["cultivo"]
                )
                fila_detalle["parcelas"] = detalle.get("parcela", "")

                if _texto(detalle.get("superficie")):

                    fila_detalle["superficie"] = detalle.get("superficie")

                filas.append(fila_detalle)

            continue

        filas.append(fila_base)

    return _dataframe(COLUMNAS_FERTILIZACION, filas)


def obtener_dataframe_practicas(conn, campana_id=None):

    practicas = _filtrar_campana(
        _leer_tabla(conn, "practicas_culturales"),
        campana_id,
    )

    if practicas.empty:

        return _dataframe(COLUMNAS_PRACTICAS)

    campanas = _leer_tabla(conn, "campanas")
    cultivos = _mapa_por_id(_leer_tabla(conn, "cultivos"))
    parcelas = _mapa_por_id(_leer_tabla(conn, "parcelas"))
    maquinaria = _mapa_por_id(_leer_tabla(conn, "maquinaria"))
    proveedores = _mapa_por_id(_leer_tabla(conn, "proveedores"))
    parcelas_por_practica = _agrupar_parcelas(
        _leer_tabla(conn, "practicas_culturales_parcelas"),
        "practica_id",
        parcelas,
    )
    detalles = _leer_detalles_actuacion(
        conn,
        "practicas_culturales_cultivos",
        "practica_id",
    )

    if not parcelas_por_practica:

        parcelas_por_practica = _agrupar_parcelas(
            _leer_tabla(conn, "practica_parcelas"),
            "practica_id",
            parcelas,
        )
    filas = []

    for _, fila in practicas.iterrows():

        practica_id = _int_o_none(fila.get("id"))
        cultivo = cultivos.get(_int_o_none(fila.get("cultivo_id")), {})
        maquinaria_item = maquinaria.get(_int_o_none(fila.get("maquinaria_id")), {})
        proveedor = proveedores.get(_int_o_none(fila.get("proveedor_id")), {})
        fila_base = {
            "practica_id": fila.get("id", ""),
            "campaña": _campana_nombre_por_id(
                campanas,
                _int_o_none(fila.get("campana_id")),
            ),
            "fecha": fila.get("fecha", ""),
            "labor": fila.get("labor", ""),
            "cultivo": _etiqueta_cultivo(cultivo) or fila.get("cultivo", ""),
            "parcelas": parcelas_por_practica.get(practica_id, ""),
            "superficie": fila.get("superficie", ""),
            "maquinaria": _descripcion_maquinaria(maquinaria_item),
            "prestador": proveedor.get("nombre", ""),
            "codigo_actuacion_siex": fila.get("codigo_actuacion_siex", ""),
            "observaciones": fila.get("observaciones", ""),
        }
        detalles_practica = (
            detalles[
                pd.to_numeric(detalles["registro_id"], errors="coerce")
                == practica_id
            ].copy()
            if not detalles.empty
            and practica_id is not None
            else pd.DataFrame()
        )

        if not detalles_practica.empty:

            for _, detalle in detalles_practica.iterrows():

                cultivo_detalle = cultivos.get(
                    _int_o_none(detalle.get("cultivo_id")),
                    cultivo,
                )
                fila_detalle = fila_base.copy()
                fila_detalle["cultivo"] = (
                    _etiqueta_cultivo(cultivo_detalle)
                    or detalle.get("cultivo", "")
                    or fila_base["cultivo"]
                )
                fila_detalle["parcelas"] = detalle.get("parcela", "")

                if _texto(detalle.get("superficie")):

                    fila_detalle["superficie"] = detalle.get("superficie")

                filas.append(fila_detalle)

            continue

        filas.append(fila_base)

    return _dataframe(COLUMNAS_PRACTICAS, filas)


def obtener_dataframe_cosecha(conn, campana_id=None):

    cosechas = _filtrar_campana(_leer_tabla(conn, "cosecha"), campana_id)

    if cosechas.empty:

        return _dataframe(COLUMNAS_COSECHA)

    campanas = _leer_tabla(conn, "campanas")
    cultivos = _mapa_por_id(_leer_tabla(conn, "cultivos"))
    clientes = _mapa_por_id(_leer_tabla(conn, "clientes"))
    parcelas = _mapa_por_id(_leer_tabla(conn, "parcelas"))
    detalles = _leer_tabla(conn, "cosecha_cultivos")
    parcelas_por_cosecha = _agrupar_parcelas(
        _leer_tabla(conn, "cosecha_parcelas"),
        "cosecha_id",
        parcelas,
    )
    filas = []

    for _, fila in cosechas.iterrows():

        cosecha_id = _int_o_none(fila.get("id"))
        cultivo = cultivos.get(_int_o_none(fila.get("cultivo_id")), {})
        cliente = clientes.get(_int_o_none(fila.get("cliente_id")), {})
        detalles_cosecha = (
            detalles[
                pd.to_numeric(detalles["cosecha_id"], errors="coerce")
                == cosecha_id
            ].copy()
            if not detalles.empty
            and "cosecha_id" in detalles.columns
            and cosecha_id is not None
            else pd.DataFrame()
        )

        if not detalles_cosecha.empty:

            for _, detalle in detalles_cosecha.iterrows():

                cultivo_detalle = cultivos.get(
                    _int_o_none(detalle.get("cultivo_id")),
                    cultivo,
                )
                parcela_detalle = parcelas.get(
                    _int_o_none(detalle.get("parcela_id")),
                    {},
                )
                filas.append({
                    "cosecha_id": fila.get("id", ""),
                    "campaña": _campana_nombre_por_id(
                        campanas,
                        _int_o_none(fila.get("campana_id")),
                    ),
                    "fecha": fila.get("fecha", ""),
                    "cultivo": (
                        _etiqueta_cultivo(cultivo_detalle)
                        or fila.get("cultivo", "")
                    ),
                    "parcelas": _etiqueta_parcela(parcela_detalle),
                    "superficie": detalle.get("superficie", ""),
                    "cantidad": fila.get("cantidad", "") or fila.get("kg", ""),
                    "unidad": (
                        fila.get("unidad", "")
                        or ("kg" if _texto(fila.get("kg")) else "")
                    ),
                    "destino": fila.get("destino", ""),
                    "cliente": cliente.get("nombre", "") or fila.get("cliente", ""),
                    "observaciones": fila.get("observaciones", ""),
                })

            continue

        parcelas_texto = parcelas_por_cosecha.get(cosecha_id, "") or fila.get(
            "parcelas",
            "",
        )
        filas.append({
            "cosecha_id": fila.get("id", ""),
            "campaña": _campana_nombre_por_id(
                campanas,
                _int_o_none(fila.get("campana_id")),
            ),
            "fecha": fila.get("fecha", ""),
            "cultivo": _etiqueta_cultivo(cultivo) or fila.get("cultivo", ""),
            "parcelas": parcelas_texto,
            "superficie": "",
            "cantidad": fila.get("cantidad", "") or fila.get("kg", ""),
            "unidad": (
                fila.get("unidad", "")
                or ("kg" if _texto(fila.get("kg")) else "")
            ),
            "destino": fila.get("destino", ""),
            "cliente": cliente.get("nombre", "") or fila.get("cliente", ""),
            "observaciones": fila.get("observaciones", ""),
        })

    return _dataframe(COLUMNAS_COSECHA, filas)


def obtener_dataframe_maquinaria(conn):

    maquinaria = _leer_tabla(conn, "maquinaria")
    equipos = _leer_tabla(conn, "equipos_aplicacion")
    filas = []

    for _, fila in maquinaria.iterrows():

        datos = fila.to_dict()
        filas.append({
            "id_visual": f"MAQ-{_texto(datos.get('id'))}",
            "origen": "Maquinaria",
            "tipo": datos.get("tipo", ""),
            "marca": datos.get("marca", ""),
            "modelo": datos.get("modelo", ""),
            "matricula": datos.get("matricula", ""),
            "numero_roma": datos.get("numero_roma", ""),
            "descripcion": _descripcion_maquinaria(datos),
            "observaciones": datos.get("observaciones", ""),
        })

    for _, fila in equipos.iterrows():

        datos = fila.to_dict()
        filas.append({
            "id_visual": f"EQ-{_texto(datos.get('id'))}",
            "origen": "Equipo aplicación",
            "tipo": datos.get("tipo", ""),
            "marca": datos.get("marca", ""),
            "modelo": datos.get("modelo", ""),
            "matricula": datos.get("matricula", ""),
            "numero_roma": datos.get("numero_roma", ""),
            "descripcion": _descripcion_maquinaria(datos),
            "observaciones": datos.get("observaciones", ""),
        })

    return _dataframe(COLUMNAS_MAQUINARIA, filas)


def obtener_dataframe_documentos(conn, campana_id=None):

    documentos = []
    tratamientos = _filtrar_campana(
        _leer_tabla(conn, "tratamientos"),
        campana_id,
    )
    movimientos = _filtrar_campana(
        _leer_tabla(conn, "movimientos_economicos"),
        campana_id,
    )
    tratamiento_ids = {
        _int_o_none(valor)
        for valor in tratamientos.get("id", pd.Series(dtype=object)).tolist()
    }
    tratamiento_ids.discard(None)
    movimiento_ids = {
        _int_o_none(valor)
        for valor in movimientos.get("id", pd.Series(dtype=object)).tolist()
    }
    movimiento_ids.discard(None)

    recetas = _leer_tabla(conn, "tratamientos_documentos")

    if not recetas.empty:

        if "tipo_documento" in recetas.columns:

            recetas = recetas[
                recetas["tipo_documento"].fillna("").astype(str) == "receta"
            ].copy()

        if campana_id is not None and "tratamiento_id" in recetas.columns:

            recetas = recetas[
                recetas["tratamiento_id"]
                .map(_int_o_none)
                .isin(tratamiento_ids)
            ].copy()

        for _, fila in recetas.iterrows():

            documentos.append({
                "documento_id": f"REC-{_texto(fila.get('id'))}",
                "tipo_documento": fila.get("tipo_documento", "receta"),
                "area": "Tratamientos",
                "registro_id": fila.get("tratamiento_id", ""),
                "nombre_original": fila.get("nombre_original", ""),
                "ruta_relativa": fila.get("ruta_relativa", ""),
                "sha256": fila.get("sha256", ""),
                "size_bytes": fila.get("size_bytes", ""),
                "observaciones": "Receta PDF inventariada; no incluida en el Excel.",
            })

    facturas = _leer_tabla(conn, "movimientos_economicos_documentos")

    if not facturas.empty:

        if "tipo_documento" in facturas.columns:

            facturas = facturas[
                facturas["tipo_documento"].fillna("").astype(str) == "factura"
            ].copy()

        if campana_id is not None and "movimiento_id" in facturas.columns:

            facturas = facturas[
                facturas["movimiento_id"]
                .map(_int_o_none)
                .isin(movimiento_ids)
            ].copy()

        for _, fila in facturas.iterrows():

            documentos.append({
                "documento_id": f"FAC-{_texto(fila.get('id'))}",
                "tipo_documento": fila.get("tipo_documento", "factura"),
                "area": "Contabilidad",
                "registro_id": fila.get("movimiento_id", ""),
                "nombre_original": fila.get("nombre_original", ""),
                "ruta_relativa": fila.get("ruta_relativa", ""),
                "sha256": fila.get("sha256", ""),
                "size_bytes": fila.get("size_bytes", ""),
                "observaciones": "Factura PDF inventariada; no incluida en el Excel.",
            })

    return _dataframe(COLUMNAS_DOCUMENTOS, documentos)


def obtener_dataframe_resumen(conn, campana_id, hojas):

    explotacion = hojas["Explotación"]
    campana_nombre = _campana_nombre(conn, campana_id)
    explotacion_nombre = ""
    titular = ""
    nif = ""

    if not explotacion.empty:

        primera = explotacion.iloc[0]
        explotacion_nombre = primera.get("nombre_explotacion", "")
        titular = primera.get("titular", "")
        nif = primera.get("nif", "")

    filas = [
        ("fecha_exportacion", datetime.now().isoformat(timespec="seconds")),
        ("campaña seleccionada", campana_nombre),
        ("explotación", explotacion_nombre),
        ("titular", titular),
        ("NIF", nif),
        ("número de parcelas", len(hojas["Parcelas_SIGPAC"])),
        ("número de cultivos", len(hojas["Cultivos"])),
        ("número de tratamientos", len(hojas["Tratamientos"])),
        ("número de fertilizaciones", len(hojas["Fertilización"])),
        ("número de prácticas culturales", len(hojas["Prácticas_Culturales"])),
        ("número de cosechas", len(hojas["Cosecha"])),
        ("número de documentos anexos", len(hojas["Documentos"])),
        ("versión", APP_NAME),
        ("aviso", AVISO_NO_OFICIAL),
    ]
    return pd.DataFrame(filas, columns=["campo", "valor"])


def _ajustar_hoja(hoja):

    hoja.freeze_panes = "A2"
    cabecera = PatternFill("solid", fgColor="EAF1E5")

    for celda in hoja[1]:

        celda.font = Font(bold=True)
        celda.fill = cabecera
        celda.alignment = Alignment(wrap_text=True, vertical="top")

    for fila in hoja.iter_rows():

        for celda in fila:

            celda.alignment = Alignment(wrap_text=True, vertical="top")

            if isinstance(celda.value, float):

                celda.number_format = "#,##0.00"

    for columna_celdas in hoja.columns:

        ancho = min(
            max(
                len(str(celda.value)) if celda.value is not None else 0
                for celda in columna_celdas
            ) + 2,
            55,
        )
        hoja.column_dimensions[columna_celdas[0].column_letter].width = max(
            ancho,
            10,
        )


def generar_excel_asistido_siex(campana_id=None, revision=None):

    conn = conectar()

    try:

        campana_nombre = _campana_nombre(conn, campana_id)
        validacion = (
            revision.copy()
            if revision is not None
            else pd.DataFrame(columns=COLUMNAS_VALIDACION)
        )

        for columna in COLUMNAS_VALIDACION:

            if columna not in validacion.columns:

                validacion[columna] = ""

        validacion = validacion[COLUMNAS_VALIDACION].copy()
        hojas = {
            "Validación": validacion,
            "Explotación": obtener_dataframe_explotacion(conn),
            "Parcelas_SIGPAC": obtener_dataframe_parcelas(
                conn,
                campana_nombre,
            ),
            "Cultivos": obtener_dataframe_cultivos(conn, campana_id),
            "Tratamientos": obtener_dataframe_tratamientos(conn, campana_id),
            "Fertilización": obtener_dataframe_fertilizacion(conn, campana_id),
            "Prácticas_Culturales": obtener_dataframe_practicas(
                conn,
                campana_id,
            ),
            "Cosecha": obtener_dataframe_cosecha(conn, campana_id),
            "Maquinaria": obtener_dataframe_maquinaria(conn),
            "Documentos": obtener_dataframe_documentos(conn, campana_id),
        }
        hojas = {
            "Resumen": obtener_dataframe_resumen(conn, campana_id, hojas),
            **hojas,
        }

    finally:

        conn.close()

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        for nombre_hoja, dataframe in hojas.items():

            _preparar_excel(dataframe).to_excel(
                writer,
                sheet_name=nombre_hoja,
                index=False,
            )

        for hoja in writer.sheets.values():

            _ajustar_hoja(hoja)

    buffer.seek(0)
    nombre_campana = _sanear_nombre_archivo(campana_nombre)
    nombre_archivo = (
        "cuadernopro_exportacion_asistida_siex_"
        f"{nombre_campana}_{date.today().isoformat()}.xlsx"
    )
    return buffer.getvalue(), nombre_archivo
